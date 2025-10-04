import re
import os
import PyPDF2
import arabic_reshaper
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


def fix_arabic(text: str) -> str:
    """إصلاح وتنسيق النص العربي"""
    try:
        reshaped = arabic_reshaper.reshape(text)
        return reshaped
    except Exception:
        return text


def read_pdf(path: str) -> list:
    """قراءة محتوى ملف PDF"""
    context = []
    try:
        with open(path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    context.extend(text.splitlines())
        print(f"✅ تم قراءة الملف: {os.path.basename(path)}")
    except Exception as e:
        print(f"❌ خطأ في قراءة الملف {os.path.basename(path)}: {e}")
    return context


def filters(text: list[str], filename: str) -> dict:
    """استخراج البيانات من النص"""
    due_date: list = []
    end_of_payments: list = []
    amount: list = []
    row = {"filename": filename}

    for i in range(len(text)):
        # استخراج رقم العقد
        if "Contract No" in text[i]:
            contract_no = text[i].split("Contract No")[1].split(":")[0].replace(". ", "").strip()
            row["Contract No"] = contract_no

        # استخراج تاريخ بداية العقد
        if "Tenancy Start Date" in text[i]:
            tenancy_start_date = text[i].split("Tenancy Start Date")[1].split(":")[0].strip()
            row["Tenancy Start Date"] = tenancy_start_date

        # استخراج تاريخ نهاية العقد
        if "Tenancy End Date" in text[i]:
            tenancy_end_date = text[i].split("Tenancy End Date")[1].split(":")[0].strip()
            row["Tenancy End Date"] = tenancy_end_date

        # استخراج اسم المستأجر
        if "name/Founder" in text[i]:
            company_name = "".join(text[i:i + 2]).split("Organization")[0]
            company_name = company_name.rsplit(" ", 1)[0].replace("name/Founder", "")[:-3].strip()
            row["Tenancy Name"] = fix_arabic(company_name)

        # استخراج العنوان الوطني
        if "National Address" in text[i]:
            national_address = text[i].replace("National Address", "").strip()
            row["National Address"] = national_address

        # استخراج اسم المؤجر
        if "Lessor Data" in text[i]:
            try:
                lessor_text = " ".join(text[i:i + 3])
                lessor_text = re.sub(r"\s+", " ", lessor_text).strip()

                if "Name" in lessor_text:
                    name_part = lessor_text.split("Name", 1)[1].strip()
                    name_clean = re.sub(r"^[:\s]*الاسم[:\s]*", "", name_part).strip()
                    name_clean = re.split(r"Nationality|:", name_clean)[0].strip()
                    name_clean = re.sub(r"(اﻻﺳﻢ|الاسم|\sالاسم|\sاﻻﺳﻢ)$", "", name_clean, flags=re.UNICODE).strip()

                    row["Lessor Name"] = fix_arabic(name_clean)
                else:
                    row["Lessor Name"] = ""
            except Exception as e:
                print(f"⚠️ خطأ في استخراج اسم المؤجر من {filename}: {e}")
                row["Lessor Name"] = ""

        # استخراج بيانات الدفعات
        pattern = r"^\d+\.\d+\s+\d{4}-\d{2}-\d{2}\s+\d{4}-\d{2}-\d{2}.*\d{4}-\d{2}-\d{2}\s+\d{4}-\d{2}-\d{2}\s+\d+\s*$"
        payments = re.findall(pattern, text[i])
        if payments:
            payments = "".join(payments).split()
            if len(payments) >= 6:
                due_date.append(payments[5])
                end_of_payments.append(payments[4])
                amount.append(payments[0])

    row["Due Date"] = due_date[:]
    row["End of Payments"] = end_of_payments[:]
    row["Amount"] = amount[:]

    # التحقق من وجود بيانات
    if not row.get("Contract No"):
        print(f"⚠️ لم يتم العثور على رقم العقد في الملف: {filename}")

    return row


def process_all_pdfs(folder_path: str) -> list:
    """معالجة جميع ملفات PDF في المجلد"""
    all_data = []
    if not os.path.exists(folder_path):
        print(f"❌ المجلد غير موجود: {folder_path}")
        return all_data

    pdf_files = [f for f in os.listdir(folder_path) if f.lower().endswith('.pdf')]

    if not pdf_files:
        print(f"❌ لا توجد ملفات PDF في المجلد: {folder_path}")
        return all_data

    print(f"\n📂 تم العثور على {len(pdf_files)} ملف PDF\n")

    for pdf_file in pdf_files:
        pdf_path = os.path.join(folder_path, pdf_file)
        print(f"🔄 معالجة: {pdf_file}")

        extracting = read_pdf(pdf_path)
        if extracting:
            pdf_data = filters(extracting, pdf_file)
            all_data.append(pdf_data)
            print(f"✅ تم استخراج البيانات من: {pdf_file}\n")
        else:
            print(f"⚠️ لم يتم استخراج أي نص من: {pdf_file}\n")

    return all_data


def convert_to_excel(all_data: list, output_file: str) -> None:
    """تحويل جميع البيانات إلى ملف Excel واحد"""
    if not all_data:
        print("❌ لا توجد بيانات لحفظها في Excel")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Contract Data"

    # إعداد العناوين
    headers = [
        "اسم الملف",
        "Contract No",
        "Tenancy Start Date",
        "Tenancy End Date",
        "Tenancy Name",
        "National Address",
        "Lessor Name",
        "Due Date",
        "End of Payments",
        "Amount"
    ]

    # تنسيق صف العناوين
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    current_row = 2

    # إضافة البيانات لكل عقد
    for data in all_data:
        max_payments = len(data.get('Due Date', []))

        if max_payments == 0:
            # إضافة صف واحد حتى لو لم توجد دفعات
            ws.append([
                data.get("filename", ""),
                data.get("Contract No", ""),
                data.get("Tenancy Start Date", ""),
                data.get("Tenancy End Date", ""),
                data.get("Tenancy Name", ""),
                data.get("National Address", ""),
                data.get("Lessor Name", ""),
                "", "", ""
            ])
            current_row += 1
        else:
            # إضافة صف لكل دفعة
            for idx, (due, end, amount) in enumerate(zip(
                    data.get('Due Date', []),
                    data.get('End of Payments', []),
                    data.get('Amount', [])
            )):
                if idx == 0:
                    # الصف الأول يحتوي على جميع البيانات الأساسية
                    ws.append([
                        data.get("filename", ""),
                        data.get("Contract No", ""),
                        data.get("Tenancy Start Date", ""),
                        data.get("Tenancy End Date", ""),
                        data.get("Tenancy Name", ""),
                        data.get("National Address", ""),
                        data.get("Lessor Name", ""),
                        due,
                        end,
                        amount
                    ])
                else:
                    # الصفوف التالية تحتوي فقط على بيانات الدفعات
                    ws.append(["", "", "", "", "", "", "", due, end, amount])
                current_row += 1

    # ضبط عرض الأعمدة تلقائياً
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length

        adjusted_width = min(max_length + 3, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # حفظ الملف
    wb.save(output_file)
    print(f"\n✅ تم حفظ البيانات بنجاح في: {output_file}")
    print(f"📊 عدد العقود المعالجة: {len(all_data)}")


# ============== التنفيذ الرئيسي ==============

# مسار المجلد الذي يحتوي على ملفات PDF
folder_path = r"C:\Users\ream8\Desktop\Project"
# مسار ملف Excel الناتج
excel_path = r"C:\Users\ream8\Desktop\Project\Tenant_Info.xlsx"

print("=" * 60)
print("🚀 بدء معالجة ملفات العقود")
print("=" * 60)

# معالجة جميع ملفات PDF
all_contracts_data = process_all_pdfs(folder_path)

# تحويل البيانات إلى Excel
convert_to_excel(all_contracts_data, excel_path)

print("=" * 60)
print("✅ اكتملت العملية بنجاح!")
print("=" * 60)