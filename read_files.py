import re
import PyPDF2
import arabic_reshaper
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def fix_arabic(text: str) -> str:
    """
    Fix arabic fonts
    :param text (str) : text to fixed
    :return: version of Arabic text
    """
    try:
        reshaped = arabic_reshaper.reshape(text)
        return arabic_reshaper.reshape(reshaped)
    except Exception as error:
        return text


# new structure
def read_context(path: str) -> list:
    """read all context from pdf file.
    Arguments:
        :param pdf_file (str) : path of the file.
    Returns:
        :return : return a list content all context.
    """
    # extracting contexts from pdf_file
    context = []
    with open(path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        for page in reader.pages:
            text = page.extract_text()
            if text:
                context.extend(text.splitlines())
    return context


def filters(text: list[str]) -> dict:
    """filter all text
    Arguments:
        :param text (str) : text to find the match
    Returns:
        :return : return filtered text
    """
    # Values
    due_date:list = []
    end_of_payments: list = []
    amount: list = []
    # row data
    row = {}

    for i in range(len(text)):
        # Contract No
        if "Contract No" in text[i]:
            # filter contract no
            contract_no = text[i].split("Contract No")[1].split(":")[0].replace(". ", "")
            # add to data row
            row["Contract No"] = contract_no
        # Tenancy Start Date
        if "Tenancy Start Date" in text[i]:
            # filter tenancy start date
            tenancy_start_date = text[i].split("Tenancy Start Date")[1].split(":")[0]
            # add to data row
            row["Tenancy Start Date"] = tenancy_start_date
        # Tenancy End Date
        if "Tenancy End Date" in text[i]:
            # filter tenancy end date
            tenancy_end_date = text[i].split("Tenancy End Date")[1].split(":")[0]
            # add to data row
            row["Tenancy End Date"] = tenancy_end_date
        # Company Name
        if "name/Founder" in text[i]:
            # filter company name
            company_name = "".join(text[i:i + 2]).split("Organization")[0]
            # filter only company name without "name/Founder"
            company_name = company_name.rsplit(" ", 1)[0].replace("name/Founder", "")[:-3]
            # add to data row with fix Arabic font
            row["Tenancy Name"] = fix_arabic(company_name)
            
        # National Address
        if "National Address" in text[i]:
            # filter national address
            national_address = text[i].replace("National Address", "")
            # add to data row
            row["National Address"] = national_address

        # matches
        pattern = r"^\d+\.\d+\s+\d{4}-\d{2}-\d{2}\s+\d{4}-\d{2}-\d{2}.*\d{4}-\d{2}-\d{2}\s+\d{4}-\d{2}-\d{2}\s+\d+\s*$"
        # match all text with pattern 
        payments = re.findall(pattern, text[i])
        # Payments schedular
        if payments:
            # split data
            payments = "".join(payments).split(" ")
            # add all due date of contract as list to filter
            due_date.append(payments[5])
            # add all end of payments of contract as list to filter
            end_of_payments.append(payments[4])
            # add all amounts of contract as list to filter
            amount.append(payments[0])

    # add values to data row
    row["Due Date"] = due_date[:]
    # add values to data row
    row["End of Payments"] = end_of_payments[:]
    # add values to data row
    row["Amount"] = amount[:]
    
    return row


def write_to_excel(data, output_file: str) -> None:
    """convert all data to excel.
    Arguments:
    :param data (dict) : data to write in DB
    :param output_file (str) : file name to save DB
    :return : save file or append data to execute
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Contract Data"

    # Write general contract info at the top
    ws.append(list(data.keys())[:])
    ws.append(list(data.values())[:5])
    # counter for multipule data in rows
    counter = 2
    for due, end, amount in zip(data['Due Date'], data['End of Payments'], data['Amount']):
        ws[f"F{counter}"] = due
        ws[f"G{counter}"] = end
        ws[f"H{counter}"] = amount
        counter += 1

    column_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                # Get column index and convert to letter
                column_letter = get_column_letter(cell.column)
                # Calculate length of the value
                cell_length = len(str(cell.value))
                # Update the max length for the column
                if column_letter not in column_widths:
                    column_widths[column_letter] = cell_length
                else:
                    if cell_length > column_widths[column_letter]:
                        column_widths[column_letter] = cell_length

    # Set column widths
    for col_letter, col_width in column_widths.items():
        ws.column_dimensions[col_letter].width = col_width + 2

    # file name to save DB
    wb.save(output_file)

if __name__=="__main__":
    # path of execute file to read
    pdf_context =  r"C:\Users\ream8\Desktop\Project\10988496532.pdf"
    # use read_context method
    extracting = read_context(pdf_context)
    # use filters method
    pdf_data = filters(extracting)
    # File name of Excel save in executed floder
    excel_path = r"10988496532.xlsx"
    # use convert_to_excel method to write and save file
    convert_to_excel(pdf_data, excel_path)

